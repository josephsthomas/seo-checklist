import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json, sys

def write_role_report(role_number, role_name, defects, output_path):
    """Write defects list to a formatted Excel file.
    defects: list of dicts with keys:
        bug_id, severity, category, component, file_line,
        description, steps, expected, actual, impact
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Role {role_number} - {role_name}"

    # Headers
    headers = ["Bug ID", "Role", "Severity", "Category", "Component",
               "File:Line", "Description", "Steps to Reproduce",
               "Expected Behavior", "Actual Behavior", "Impact", "Status"]

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Severity colors
    severity_fills = {
        "CRITICAL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "HIGH": PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
        "LOW": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    }
    severity_fonts = {
        "CRITICAL": Font(color="FFFFFF", bold=True),
        "HIGH": Font(color="FFFFFF", bold=True),
        "MEDIUM": Font(color="000000", bold=True),
        "LOW": Font(color="000000"),
    }

    # Helper to convert lists/non-strings to strings for Excel
    def to_str(val):
        if isinstance(val, list):
            return "\n".join(str(v) for v in val)
        if val is None:
            return ""
        return str(val)

    # Data rows
    for i, defect in enumerate(defects, 2):
        ws.cell(row=i, column=1, value=to_str(defect.get("bug_id", ""))).border = thin_border
        ws.cell(row=i, column=2, value=role_name).border = thin_border

        sev = to_str(defect.get("severity", "MEDIUM"))
        sev_cell = ws.cell(row=i, column=3, value=sev)
        sev_cell.fill = severity_fills.get(sev, PatternFill())
        sev_cell.font = severity_fonts.get(sev, Font())
        sev_cell.border = thin_border
        sev_cell.alignment = Alignment(horizontal='center')

        ws.cell(row=i, column=4, value=to_str(defect.get("category", ""))).border = thin_border
        ws.cell(row=i, column=5, value=to_str(defect.get("component", ""))).border = thin_border
        ws.cell(row=i, column=6, value=to_str(defect.get("file_line", ""))).border = thin_border
        ws.cell(row=i, column=7, value=to_str(defect.get("description", ""))).border = thin_border
        ws.cell(row=i, column=8, value=to_str(defect.get("steps", ""))).border = thin_border
        ws.cell(row=i, column=9, value=to_str(defect.get("expected", ""))).border = thin_border
        ws.cell(row=i, column=10, value=to_str(defect.get("actual", ""))).border = thin_border
        ws.cell(row=i, column=11, value=to_str(defect.get("impact", ""))).border = thin_border
        ws.cell(row=i, column=12, value="OPEN").border = thin_border

        # Wrap text for long fields
        for col in [7, 8, 9, 10, 11]:
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True)

    # Column widths
    widths = [10, 25, 12, 22, 20, 50, 50, 50, 40, 40, 35, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(defects) + 1}"

    wb.save(output_path)
    print(f"SUCCESS: Wrote {len(defects)} defects to {output_path}")
    return len(defects)

if __name__ == "__main__":
    # Usage: python write_defects.py <json_file> <role_number> <role_name> <output_path>
    with open(sys.argv[1]) as f:
        defects = json.load(f)
    write_role_report(int(sys.argv[2]), sys.argv[3], defects, sys.argv[4])
